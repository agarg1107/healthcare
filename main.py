from tkinter import *
from datetime import date
from tkinter import messagebox
from PIL import Image
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib
import tkinter as tk
import customtkinter

import dialogbox

customtkinter.set_appearance_mode("System")
textcolor = "#333333"
side_frame_col = "#F2DFD7"
buttoncolor = "#FF5722"
buttoncolorlite = "#FF855F"
background = "#D4C1EC"
obj_frame_col = "#9F9FED"
mainbackground = "#736ced"
framebg = "#EDEDED"
framefg = "#06283D"
root = Tk()
maincol1 = "pink"
maincol2 = "pink"
maincol3 = "pink"
width = root.winfo_screenwidth()
height = root.winfo_screenheight()

root.title("Clinic Managment System")
root.state('zoomed')
# root.geometry ("%dx%d"%(width,height))
root.config(bg=mainbackground)
# pathmain = "//LAPTOP-F1A0LRP8/Users/aman/Student_data.xlsx"
pathmain = "Student_data.xlsx"
file = pathlib.Path(pathmain)

if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active

    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "D0B"
    sheet['F1'] = "Date Of Registration"
    sheet['G1'] = "weight"
    sheet['H1'] = "height"
    sheet['I1'] = "Temperature"
    sheet['J1'] = "Respiration"
    sheet['K1'] = "Pulse"
    sheet['L1'] = "BP"
    sheet['M1'] = "Village"
    sheet['N1'] = "Mobile"

    file.save(pathmain)


def update():
    root.after(1000, update)  # run itself again after 1000 ms


def Exit():
    root.destroy()


def registration_no():
    file = openpyxl.load_workbook(pathmain)
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)

    except:
        Registration.set("1")


def Clear():
    global img
    Name.set('')
    age.set('')
    weight.set('')
    height.set('')
    temprature.set('')
    respiration.set('')
    pulse.set('')
    bp.set('')
    village.set('')
    mobile.set('')

    registration_no()


def Clear2():
    saveButton.configure(state='active')
    global img
    Name.set('')
    age.set('')
    weight.set('')
    height.set('')
    village.set('')
    mobile.set('')
    temprature.set('')
    respiration.set('')
    pulse.set('')
    bp.set('')

    registration_no()


def Save():
    R1 = Registration.get()
    N1 = Name.get()

    try:
        G1 = gender
    except:
        messagebox.showerror("“error", "Select Gender!")

    D2 = age.get()
    D1 = Date.get()
    Re1 = weight.get()
    S1 = height.get()
    temp = temprature.get()
    res = respiration.get()
    F1 = pulse.get()
    M1 = bp.get()
    vill = village.get()
    mob = mobile.get()
    mobile.set('')
    if N1 == "" or D2 == "" or Re1 == "" or S1 == "" or temp == "" or res == "" or F1 == "" or M1 == "":
        messagebox.showerror("“error", "Few Data is missing!")
    else:
        file = openpyxl.load_workbook(pathmain)
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)

        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=Re1)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=temp)
        sheet.cell(column=10, row=sheet.max_row, value=res)
        sheet.cell(column=11, row=sheet.max_row, value=F1)
        sheet.cell(column=12, row=sheet.max_row, value=M1)
        sheet.cell(column=13, row=sheet.max_row, value=vill)
        sheet.cell(column=14, row=sheet.max_row, value=mob)

        file.save(pathmain)
        try:
            img.save("Student Images/" + str(R1) + ".jpg")
        except:
            messagebox.showinfo("info", "Profile Picture is not available!!!!")
        messagebox.showinfo("info", "Sucessfully data entered!!!")
        Clear()  # clear entry box and image section

        registration_no()  # it will recheck registration no. and reissue new no.


first =0
def search2():
    text = Search.get()  # taking input from entry box

    Clear()  # to clear all the data already available in entry box and other
    saveButton.configure(
        state='disable')  # after clicking on search , save button will disable so that no one can click on

    file = openpyxl.load_workbook(pathmain)
    sheet = file.active
    if(text == ""):
        messagebox.showerror("Invalid", "Please enter mobile number")
        return
    for row in sheet.rows:
        if row[12].value == int(text):
            name = row[12]
            reg_no_position = str(name)[14:-1]
            global first
            first =1;
            reg_number = str(name)[15:-1]
            try:
                print(str(name))
            except:
                messagebox.showerror("Invalid", "Invalid registration number! !!")
    if(first == 0):
        messagebox.showerror("Invalid", "Invalid registration number! !!")
        return
    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value
    x13 = sheet.cell(row=int(reg_number), column=13).value
    x14 = sheet.cell(row=int(reg_number), column=14).value

    Registration.set(x1)
    Name.set(x2)


    if x4 == "Female":
        R2.select()

    else:
        R1.select()

    age.set(x5)

    Date.set(x6)

    weight.set(x7)

    height.set(x8)
    temprature.set(x9)

    respiration.set(x10)
    pulse.set(x11)
    bp.set(x12)
    village.set(x13)
    mobile.set(x14)
    first =0;
def search():
    text = Search.get()  # taking input from entry box

    Clear()  # to clear all the data already available in entry box and other
    saveButton.configure(
        state='disable')  # after clicking on search , save button will disable so that no one can click on

    file = openpyxl.load_workbook(pathmain)
    sheet = file.active
    if(text == ""):
        messagebox.showerror("Invalid", "Invalid registration number! !!")
        return
    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            reg_no_position = str(name)[14:-1]
            global first
            first =1;
            reg_number = str(name)[15:-1]
            try:
                print(str(name))
            except:
                messagebox.showerror("Invalid", "Invalid registration number! !!")
    if(first == 0):
        messagebox.showerror("Invalid", "Invalid registration number! !!")
        return
    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value
    x13 = sheet.cell(row=int(reg_number), column=13).value
    x14 = sheet.cell(row=int(reg_number), column=14).value

    Registration.set(x1)
    Name.set(x2)


    if x4 == "Female":
        R2.select()

    else:
        R1.select()

    age.set(x5)

    Date.set(x6)

    weight.set(x7)

    height.set(x8)
    temprature.set(x9)

    respiration.set(x10)
    pulse.set(x11)
    bp.set(x12)
    village.set(x13)
    mobile.set(x14)
    first =0;
    ####################################Update#####################################


def Update():
    reg_number = IntVar()
    R1 = Registration.get()
    N1 = Name.get()

    selection()
    Gl = gender
    D2 = age.get()
    D1 = Date.get()
    Re1 = weight.get()
    S1 = height.get()
    fathername = temprature.get()
    mothername = respiration.get()
    F1 = pulse.get()
    M1 = bp.get()
    uvill = village.get()
    umob = mobile.get()
    file = openpyxl.load_workbook(pathmain)
    sheet = file.active
    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            print(reg_number)

    sheet.cell(column=1, row=int(reg_number), value=R1)
    sheet.cell(column=2, row=int(reg_number), value=N1)

    sheet.cell(column=4, row=int(reg_number), value=Gl)
    sheet.cell(column=5, row=int(reg_number), value=D2)
    sheet.cell(column=6, row=int(reg_number), value=D1)
    sheet.cell(column=7, row=int(reg_number), value=Re1)
    sheet.cell(column=8, row=int(reg_number), value=S1)
    sheet.cell(column=9, row=int(reg_number), value=fathername)
    sheet.cell(column=10, row=int(reg_number), value=mothername)
    sheet.cell(column=11, row=int(reg_number), value=F1)
    sheet.cell(column=12, row=int(reg_number), value=M1)
    sheet.cell(column=13, row=int(reg_number), value=uvill)
    sheet.cell(column=14, row=int(reg_number), value=umob)
    file.save(pathmain)
    messagebox.showinfo("info", "Sucessfully data entered!!!")
    Clear()
    ####################################gender#####################################


def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"


fontmain = "Dotum"
Name = StringVar()
age = StringVar()
radio = IntVar()
weight = StringVar()
height = StringVar()
temprature = StringVar()
pulse = StringVar()
respiration = StringVar()
bp = StringVar()
obj = None
mobile = StringVar()
village = StringVar()
Search = StringVar()

# top frames
obj = customtkinter.CTkFrame(master=root, corner_radius=15, width=900, height=500, fg_color=obj_frame_col, border_width=4,
                             border_color="black")
obj.place(x=300, y=200)

Label(root, text="Clinic Management", width=10, height=2, bg="#c36464", fg='#fff', font='arial 20 bold').pack(side=TOP,
                                                                                                              fill=X)

# Registration and Date

customtkinter.CTkLabel(master=root, text="Token No:", text_color=textcolor,font=(fontmain, 20)).place(x=300, y=157)
customtkinter.CTkLabel(master=root, text="Date:", text_color=textcolor,font=(fontmain, 20)).place(x=910, y=157)
Registration = IntVar()
Date = StringVar()
reg_entry = customtkinter.CTkEntry(master=root, corner_radius=15, fg_color=background,text_color=textcolor,textvariable=Registration, height=40,
                                   font=(fontmain, 20), width=220)
reg_entry.place(x=400, y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = customtkinter.CTkEntry(master=root, corner_radius=15,fg_color=background,text_color=textcolor, textvariable=Date, height=40, font=(fontmain, 20),
                                    width=220)
date_entry.place(x=970, y=150)
Date.set(d1)


# frame
def reg_page():
    # Labels
    customtkinter.CTkLabel(obj, text="Full Name:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=55)
    customtkinter.CTkLabel(obj, text="Age:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=105)
    customtkinter.CTkLabel(obj, text="Gender:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=155)

    customtkinter.CTkLabel(obj, text="Weight:", text_color=textcolor,font=(fontmain, 20)).place(x=480, y=105)
    customtkinter.CTkLabel(obj, text="Height:", text_color=textcolor,font=(fontmain, 20)).place(x=480, y=155)
    customtkinter.CTkLabel(obj, text="Temperature:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=205)
    customtkinter.CTkLabel(obj, text="Pulse:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=255)
    customtkinter.CTkLabel(obj, text="Respiration:", text_color=textcolor,font=(fontmain, 20)).place(x=480, y=205)
    customtkinter.CTkLabel(obj, text="BP:", text_color=textcolor,font=(fontmain, 20)).place(x=480, y=255)
    customtkinter.CTkLabel(obj, text="Village Name:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=305)
    customtkinter.CTkLabel(obj, text="Mobile:", text_color=textcolor,font=(fontmain, 20)).place(x=480, y=305)

    # Entry
    name_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=15, textvariable=Name, height=40, font=(fontmain, 20),
                                        width=220)
    name_entry.place(x=170, y=50)
    age_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=15, textvariable=age, height=40, font=(fontmain, 20),
                                       width=220)
    age_entry.place(x=170, y=100)
    weight_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=15, textvariable=weight, height=40,
                                          font=(fontmain, 20), width=220)
    weight_entry.place(x=630, y=100)
    height_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=15, textvariable=height, height=40,
                                          font=(fontmain, 20), width=220)
    height_entry.place(x=630, y=150)
    temp_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=15, textvariable=temprature, height=40,
                                        font=(fontmain, 20), width=220)
    temp_entry.place(x=170, y=200)
    pulse_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=15, textvariable=pulse, height=40,
                                         font=(fontmain, 20), width=220)
    pulse_entry.place(x=170, y=250)
    resp_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=15, textvariable=respiration, height=40,
                                        font=(fontmain, 20), width=220)
    resp_entry.place(x=630, y=200)
    bp_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=15, textvariable=bp, height=40, font=(fontmain, 20),
                                      width=220)
    bp_entry.place(x=630, y=250)
    village_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=15, textvariable=village, height=40,
                                           font=(fontmain, 20), width=220)
    village_entry.place(x=630, y=300)
    mobile_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=15, textvariable=mobile, height=40,
                                          font=(fontmain, 20), width=220)
    mobile_entry.place(x=170, y=300)

    # Radio Button
    global R1
    global R2
    R1 = customtkinter.CTkRadioButton(obj, text="Male", command=selection, variable=radio, value=1, text_color="black")
    R2 = customtkinter.CTkRadioButton(obj, text="Female", command=selection, variable=radio, value=2,
                                      text_color='black')


    R1.place(x=170, y=160)
    R2.place(x=270, y=160)



def bill_page():
    Label(obj, text="Full Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)


def stock_page():
    s_p = tk.Frame(obj)
    tk.Label(s_p, text='stock Page\n\nPage:3', font='Bold,30').place(x=1,y=1)


def del_page():
    for frame in obj.winfo_children():
        frame.destroy()


def hideindicate():
    reg_indicate.config(bg=buttoncolorlite)
    bill_indicate.config(bg=buttoncolorlite)
    stock_indicate.config(bg=buttoncolorlite)

    reg_btn.configure(fg_color=buttoncolorlite)
    bill_btn.configure(fg_color=buttoncolorlite)
    stock_btn.configure(fg_color=buttoncolorlite)


def indicate(lb, btn, page):
    hideindicate()
    if btn == 1:
        reg_btn.configure(fg_color=buttoncolor)
    elif btn == 2:
        bill_btn.configure(fg_color=buttoncolor)
    else:
        stock_btn.configure(fg_color=buttoncolor)
    lb.config(bg=buttoncolor)
    del_page()
    page()


option_frame = customtkinter.CTkFrame(master=root, corner_radius=0, fg_color=side_frame_col)
my_image = customtkinter.CTkImage(light_image=Image.open("stock-removebg-preview.png"),
                                  dark_image=Image.open("stock-removebg-preview.png"),
                                  size=(40, 40))
reg_btn = customtkinter.CTkButton(option_frame, image=my_image, fg_color=buttoncolorlite, hover="disable", text='Registration',
                                  width=150, corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                                  height=40, command=lambda: indicate(reg_indicate, 1, reg_page))
reg_btn.place(x=15, y=50)

reg_indicate = tk.Label(option_frame, text='', bg=buttoncolor)
reg_indicate.place(x=3, y=55, width=5, height=40)

bill_btn = customtkinter.CTkButton(option_frame, text="Bill                 ", hover="disable", image=my_image,
                                   fg_color=buttoncolorlite, width=150, corner_radius=10, border_width=2, border_color="black",
                                   border_spacing=2, height=40, command=lambda: indicate(bill_indicate, 2, bill_page))
bill_btn.place(x=15, y=105)

bill_indicate = tk.Label(option_frame, text='', bg=buttoncolor)
bill_indicate.place(x=3, y=110, width=5, height=40)

stock_btn = customtkinter.CTkButton(option_frame, text='Stock            ', hover="disable", image=my_image,
                                    fg_color=buttoncolorlite, width=150, corner_radius=10, border_width=2,
                                    border_color="black", border_spacing=2, height=40,
                                    command=lambda: indicate(stock_indicate, 3, stock_page))
stock_btn.place(x=15, y=160)

stock_indicate = tk.Label(option_frame, text='', bg=buttoncolor)
stock_indicate.place(x=3, y=165, width=5, height=40)

option_frame.pack(side=tk.LEFT)
option_frame.pack_propagate(False)
option_frame.configure(width=200, height=730)
main_frame = tk.Frame(root, highlightbackground='black', highlightthickness=10)

# button


customtkinter.CTkEntry(master=root, corner_radius=15,text_color=textcolor, fg_color=background,textvariable=Search, placeholder_text="search", height=40,
                       font=(fontmain, 20), width=220).place(x=1110, y=75)
imageicon3 = PhotoImage(file="Images/search.png")
srchimage = customtkinter.CTkImage(light_image=Image.open("stock-removebg-preview.png"),
                                   dark_image=Image.open("stock-removebg-preview.png"),
                                   size=(40, 40))
Srch = customtkinter.CTkButton(root, text="Search by Token", command=search, image=srchimage, fg_color=buttoncolor, hover="disable",
                               width=150, corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                               height=40)
Srch.place(x=1350, y=70)
Srch2 = customtkinter.CTkButton(root, text="Search by Mobile", command=search2, image=srchimage, fg_color=buttoncolor, hover="disable",
                               width=150, corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                               height=40)
Srch2.place(x=1350, y=130)

customtkinter.CTkButton(root, text="Upload", image=srchimage, fg_color=buttoncolor, hover="disable", width=150,
                        corner_radius=10, border_width=2, border_color="black", border_spacing=2, height=40,
                        command=Update).place(x=1300, y=370)
saveButton = customtkinter.CTkButton(root, text="Save", image=srchimage, fg_color=buttoncolor, hover="disable", width=150,
                                     corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                                     height=40, command=Save)
saveButton.place(x=1300, y=450)
customtkinter.CTkButton(root, text="Reset", image=srchimage, fg_color=buttoncolor, hover="disable", width=150,
                        corner_radius=10, border_width=2, border_color="black", border_spacing=2, height=40,
                        command=Clear2).place(x=1300, y=530)
customtkinter.CTkButton(root, text="Exit", image=srchimage, fg_color=buttoncolor, hover="disable", width=150,
                        corner_radius=10, border_width=2, border_color="black", border_spacing=2, height=40,
                        command=Exit).place(x=1300, y=610)

root.mainloop()