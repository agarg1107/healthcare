from tkinter import *
from datetime import date
from tkinter import messagebox
from PIL import Image
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib
import tkinter as tk
import customtkinter
from tkinter import ttk
#initial
customtkinter.set_appearance_mode("System")
textcolor = "#333333"
side_frame_col = "#F2DFD7"
buttoncolor = "#FF5722"
buttoncolorlite = "#FF855F"
background = "#D4C1EC"
obj_frame_col = "#9F9FED"
mainbackground = "#736ced"
framebg = "#EDEDED"
framefg = "#06283D"
root = Tk()
maincol1 = "pink"
maincol2 = "pink"
maincol3 = "pink"
width = root.winfo_screenwidth()
height = root.winfo_screenheight()

root.title("Clinic Managment System")
root.state('zoomed')
# root.geometry ("%dx%d"%(width,height))
root.config(bg=mainbackground)
# pathmain = "//LAPTOP-F1A0LRP8/Users/aman/Student_data.xlsx"
pathmain = "Student_data.xlsx"
file = pathlib.Path(pathmain)
patient_detail_from_doc = "patient.xlsx"
doc_med_old = "docmed.xlsx"
file = pathlib.Path(patient_detail_from_doc)

# ----------------------------------------------------------------------------------------------------------------------
# Declare
fontmain = "Dotum"
Name = StringVar()
age = StringVar()
radio = IntVar()
weight = StringVar()
height = StringVar()
temprature = StringVar()
pulse = StringVar()
respiration = StringVar()
bp = StringVar()
obj = None
mobile = StringVar()
village = StringVar()
Search = StringVar()
Gender = StringVar()
my_list = Listbox
my_entry = Entry

name_lable = customtkinter.CTkLabel
med_no = IntVar()
med_name = StringVar()
pnt_dis = StringVar()
dose = StringVar()
days = StringVar()
Registration = IntVar()
Date = StringVar()

docmedslst = []
#-----------------------------------------------------------------------------------------------------------------------
# file creation if not exist

if file.exists():
    pass
else:
    file = Workbook()
    sheet1 = file.active
    file.save(patient_detail_from_doc)
file = pathlib.Path(doc_med_old)
if file.exists():
    pass
else:
    file = Workbook()
    sheet1 = file.active
    sheet1['A1'] = "Name"
    file.save(doc_med_old)
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active

    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "D0B"
    sheet['F1'] = "Date Of Registration"
    sheet['G1'] = "weight"
    sheet['H1'] = "height"
    sheet['I1'] = "Temperature"
    sheet['J1'] = "Respiration"
    sheet['K1'] = "Pulse"
    sheet['L1'] = "BP"
    sheet['M1'] = "Village"
    sheet['N1'] = "Mobile"

    file.save(pathmain)


def update():
    root.after(1000, update)  # run itself again after 1000 ms
def Exit():
    root.destroy()
# def registration_no():
#     file = openpyxl.load_workbook(pathmain)
#     sheet = file.active
#     row = sheet.max_row
#
#     max_row_value = sheet.cell(row=row, column=1).value
#
#     try:
#         Registration.set(max_row_value + 1)
#
#     except:
#         Registration.set("1")
def Clear():
    global img
    Name.set('')
    age.set('')
    weight.set('')
    Gender.set('')
    height.set('')
    temprature.set('')
    respiration.set('')
    pulse.set('')
    bp.set('')
    village.set('')
    mobile.set('')
    # registration_no()
def Clear2():

    global img
    Name.set('')
    age.set('')
    weight.set('')
    Gender.set('')
    height.set('')
    village.set('')
    mobile.set('')
    temprature.set('')
    respiration.set('')
    pulse.set('')
    bp.set('')

    # registration_no()
def Save():
    print("abcd")
def search():
    text = Search.get()  # taking input from entry box

    Clear()  # to clear all the data already available in entry box and other
  # after clicking on search , save button will disable so that no one can click on

    file = openpyxl.load_workbook(pathmain)
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            try:
                print(str(name))
            except:
                messagebox.showerror("Invalid", "Invalid registration number! !!")

    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value
    x13 = sheet.cell(row=int(reg_number), column=13).value
    x14 = sheet.cell(row=int(reg_number), column=14).value

    Registration.set(x1)
    Name.set(x2)


    if x4 == "Female":
        Gender.set("Female")

    else:
        Gender.set("male")

    age.set(x5)

    Date.set(x6)

    weight.set(x7)

    height.set(x8)
    temprature.set(x9)

    respiration.set(x10)
    pulse.set(x11)
    bp.set(x12)
    village.set(x13)
    mobile.set(x14)
    remove_all()
    data.clear()
    updatetree()
    ####################################gender#####################################
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"
data = []

def updatetree():

    file = openpyxl.load_workbook(patient_detail_from_doc)
    # sheet.cell(column=1, row=sheet.max_row + 1, value=med_no)
    sheet1 = file.active

    for row in sheet1.rows:
        lst = []
        if Search.get() == row[0].value:
            lst.append(row[0].value)
            lst.append(row[1].value)
            lst.append(row[2].value)
            lst.append(row[3].value)
            data.append(lst)

    global count
    count = 0

    for record in data:
        if count % 2 == 0:
            my_tree.insert(parent='', index='end', iid=count, text="", values=(record[0], record[1], record[2],record[3]),
                           tags=('evenrow',))
        else:
            my_tree.insert(parent='', index='end', iid=count, text="", values=(record[0], record[1], record[2],record[3]),
                           tags=('oddrow',))

        count += 1
def filldata():
    file = openpyxl.load_workbook(patient_detail_from_doc)
    # sheet.cell(column=1, row=sheet.max_row + 1, value=med_no)
    sheet5 = file.active
    setdata = StringVar()
    count = 1
    rows_to_delete = []

    for rowcheck in sheet5.rows:
        if(Search.get() == rowcheck[0].value):

            rows_to_delete.append(count)
        count = count +1
    # for item in my_tree.get_children():
    #     values = my_tree.item(item)["values"]
    #
    #     for row1 in sheet5.iter_rows():
    #         count += 1
    #
    #         print(values[0])
    #         if Search.get() == str(values[0]):
    #             rows_to_delete.append(count)

    # Delete the rows in reverse order
    print(rows_to_delete)
    for row_index in sorted(rows_to_delete, reverse=True):
        sheet5.delete_rows(row_index)


    count  = 0
    for item in my_tree.get_children():
        values = my_tree.item(item)["values"]
        row = sheet5.max_row + 1
        for col, value in enumerate(values, start=1):
            setdata.set(value)
            sheet5.cell(row=row, column=col, value=setdata.get())

    count =0


    file.save(patient_detail_from_doc)

    # Close the workbook
    file.close()



def remove_all():
	for record in my_tree.get_children():
		my_tree.delete(record)
def reg_page():

    file = openpyxl.load_workbook("Stock.xlsx")
    sheet3 = file.active
    for i in sheet3.rows:
        docmedslst.append(i[2].value)

    # Labels
    customtkinter.CTkLabel(obj, text="Full Name:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=55)
    customtkinter.CTkLabel(obj, text="Age:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=105)
    customtkinter.CTkLabel(obj, text="Gender:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=155)
    customtkinter.CTkLabel(obj, text="Weight:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=355)
    customtkinter.CTkLabel(obj, text="Height:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=405)
    customtkinter.CTkLabel(obj, text="Temperature:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=205)
    customtkinter.CTkLabel(obj, text="Pulse:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=255)
    customtkinter.CTkLabel(obj, text="Respiration:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=455)
    customtkinter.CTkLabel(obj, text="BP:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=505)
    customtkinter.CTkLabel(obj, text="Village Name:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=305)
    customtkinter.CTkLabel(obj, text="Mobile:", text_color=textcolor,font=(fontmain, 20)).place(x=30, y=555)

    # Entry
    customtkinter.CTkLabel(obj, textvariable = Name, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=50)
    customtkinter.CTkLabel(obj, textvariable = age, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=100)
    customtkinter.CTkLabel(obj, textvariable = weight, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=350)
    customtkinter.CTkLabel(obj, textvariable = height, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=400)
    customtkinter.CTkLabel(obj, textvariable = temprature, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=200)
    customtkinter.CTkLabel(obj, textvariable = pulse, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=250)
    customtkinter.CTkLabel(obj, textvariable = respiration, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=450)
    customtkinter.CTkLabel(obj, textvariable = bp, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=500)
    customtkinter.CTkLabel(obj, textvariable = village, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=550)
    customtkinter.CTkLabel(obj, textvariable = mobile, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=300)
    customtkinter.CTkLabel(obj, textvariable= Gender, text_color=textcolor, font=(fontmain, 20)).place(x=170, y=160)

    # Doc
    def clear_textbox():
        textbox.delete(1.0,3.0)
    def get_text():
        global pnt_dis
        pnt_dis = textbox.get(1.0, 3.0)

    customtkinter.CTkLabel(obj2, text="Patient Description", text_color=textcolor, font=(fontmain, 20)).place(x=430, y=10)
    textbox = customtkinter.CTkTextbox(obj2, fg_color=mainbackground,width=400,height=200, corner_radius=10, border_width=2,
                                        border_color="black", border_spacing=2,text_color="black",activate_scrollbars=True,scrollbar_button_color=background)
    textbox.place(x=430, y=50)



    def update(data):
        # Clear the listbox
        my_list.delete(0, END)

        # Add toppings to listbox
        for item in data:
            my_list.insert(END, item)
    def fillout(e):
        my_entry.delete(0, END)

        selected_item = my_list.get(ANCHOR)
        my_entry.insert(0, selected_item)

        my_list.selection_clear(0, END)
    def check(e):
        # grab what was typed
        typed = my_entry.get()

        if typed == '':
            data = toppings
        else:
            data = []
            for item in toppings:
                if typed.lower() in item.lower():
                    data.append(item)

        # update our listbox with selected items
        update(data)

    customtkinter.CTkLabel(obj2, text="Search", text_color=textcolor, font=(fontmain, 20)).place(x=10, y=10)
    my_entry = customtkinter.CTkEntry(master=obj2, text_color=textcolor,fg_color=background,corner_radius=10, textvariable=med_name, height=30,
                                        font=(fontmain, 20), width=240)
    my_entry.place(x=10, y=50)

    my_list = Listbox(obj2, width=40,bd=0,background=background)
    my_list.place(x=10, y=90)

    # Create a list of pizza toppings
    toppings = []
    file = openpyxl.load_workbook("Stock.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if(row[2].value != None):
            toppings.append(row[2].value)

    # Add the toppings to our list
    update(toppings)

    # Create a binding on the listbox onclick
    my_list.bind("<<ListboxSelect>>", fillout)

    # Create a binding on the entry box
    my_entry.bind("<KeyRelease>", check)
#     addition med
    my_entry_mg = customtkinter.CTkEntry(master=obj2, text_color=textcolor, fg_color=background, corner_radius=8,
                                      textvariable=dose, height=30,
                                      font=(fontmain, 20), width=50)
    my_entry_mg.place(x=330, y=130)
    customtkinter.CTkLabel(obj2, text="Dose : ", text_color=textcolor, font=(fontmain, 20)).place(x=260, y=130)
    my_entry_days = customtkinter.CTkEntry(master=obj2, text_color=textcolor, fg_color=background, corner_radius=8,
                                         textvariable=days, height=30,
                                         font=(fontmain, 20), width=50)
    my_entry_days.place(x=330, y=190)
    customtkinter.CTkLabel(obj2, text="Days : ", text_color=textcolor, font=(fontmain, 20)).place(x=260, y=190)
#     button delete ans add


    def Clear():
        global pnt_dis
        med_no.set('')
        days.set('')
        dose.set('')
        pnt_dis = ''
        med_name.set('')
        # registration_no()

    def add_record():
        sno =1

        for i in docmedslst:
            print(i)
            print(med_name.get())
            if(med_name.get() != i):
                print("------------------------------")
                print(i)
                print(med_name.get())
                sno = sno +1
                print("--------------------------------")
            else:
                break
        file = openpyxl.load_workbook("Stock.xlsx")
        sheet = file.active
        col = "B"
        main_drug_name = sheet[f"{col}{sno}"].value
        my_tree.tag_configure('oddrow', background="white")
        my_tree.tag_configure('evenrow', background="lightblue")

        global count
        if count % 2 == 0:
            my_tree.insert(parent='', index='end', iid=count, text="",
                           values=(Registration.get(), main_drug_name, days.get(),dose.get()), tags=('evenrow',))
        else:
            my_tree.insert(parent='', index='end', iid=count, text="",
                           values=(Registration.get(), main_drug_name, days.get(),dose.get()), tags=('oddrow',))

        count += 1

    def find_item_in_list(lst, item):
        try:
            index = lst.index(item)
            return False
        except ValueError:
            return True
    def addmed():
        if(Search.get() == ''):
            messagebox.showerror("“error", "Please find the Patient!")
            return
        get_text()
        n = med_name.get()
        med_d = days.get()
        med_dose = dose.get()

        if n == "" or med_d == "" or med_dose == "":
            messagebox.showerror("“error", "Few Data is missing!")

        else:
            # file = openpyxl.load_workbook(patient_detail_from_doc)
            # # sheet.cell(column=1, row=sheet.max_row + 1, value=med_no)
            # sheet1 = file.active
            # sheet1.cell(column=1, row=sheet1.max_row + 1, value=Search.get())
            # sheet1.cell(column=2, row=sheet1.max_row, value=n)
            # sheet1.cell(column=3, row=sheet1.max_row, value=med_d)
            # sheet1.cell(column=4, row=sheet1.max_row, value=med_dose)
            # sheet1.cell(column=5, row=sheet1.max_row, value=pnt_dis)
            #
            # file.save(patient_detail_from_doc)
            add_record()
            a = find_item_in_list(docmedslst,med_name.get())
            if(a):
                file2 = openpyxl.load_workbook(doc_med_old)

                sheet3 = file2.active
                sheet3.cell(column=1, row=sheet3.max_row + 1, value=med_name.get())
                file2.save(doc_med_old)
                for row in sheet3.rows:
                    toppings.append(row[0].value)
                    # Add the toppings to our list


            Clear() # clear entry box and image section


    def delmed():
        # x = my_tree.selection()[0]
        #
        # print(x)
        # values = my_tree.item(x, 'values')
        # print(values)
        # file3 = openpyxl.load_workbook(patient_detail_from_doc)
        # sheet4 = file3.active
        # print(values[1])
        # count =0
        # for row1 in sheet4.rows:
        #     count = count+1
        #     print(row1)
        #     if str(values[1]) == row1[1].value:
        #         sheet4.delete_rows(count)
        #         # Save the changes
        # file3.save(patient_detail_from_doc)
        # count =0
        # # Close the workbook
        # file3.close()
        # my_tree.delete(x)
        x = my_tree.selection()[0]
        my_tree.delete(x)
    add_btn = customtkinter.CTkButton(obj2, text='ADD', hover="disable",
                                        fg_color=buttoncolorlite, width=80, corner_radius=10, border_width=2,
                                        border_color="black", border_spacing=2, height=40,command=lambda: addmed()
                                      )
    add_btn.place(x=30, y=280)
    delete_btn = customtkinter.CTkButton(obj2, text='DELETE', hover="disable",
                                      fg_color=buttoncolorlite, width=100, corner_radius=10, border_width=2,
                                      border_color="black", border_spacing=2, height=40,
                                      command=lambda: delmed()
                                      )
    delete_btn.place(x=140, y=280)
#     excel

    style = ttk.Style()
    # Pick a theme
    style.theme_use("clam")
    # Configure our treeview colors

    style.configure("Treeview",
                    background=background,
                    foreground="black",
                    rowheight=20,
                    fieldbackground=mainbackground
                    )
    # Change selected color
    style.map('Treeview',
              background=[('selected', 'blue')])

    # Create Treeview Frame
    tree_frame = Frame(obj2)
    tree_frame.place(x = 20, y = 340)

    # Treeview Scrollbar
    tree_scroll = customtkinter.CTkScrollbar(tree_frame,corner_radius=9,fg_color=mainbackground,button_color=background)
    tree_scroll.pack(side=RIGHT, fill=Y)

    # Create Treeview
    global my_tree
    my_tree = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set, selectmode="extended")
    # Pack to the screen
    my_tree.pack()

    # Configure the scrollbar
    tree_scroll.configure(command=my_tree.yview)

    # Define Our Columns
    my_tree['columns'] = ("Token", "Name", "Dose","Days")

    # Formate Our Columns
    my_tree.column("#0", width=0, stretch=NO)
    my_tree.column("Name", anchor=W, width=140)
    my_tree.column("Dose", anchor=CENTER, width=100)
    my_tree.column("Days", anchor=W, width=140)
    my_tree.column("Token", anchor=W, width=140)

    # Create Headings
    my_tree.heading("#0", text="", anchor=W)
    my_tree.heading("Name", text="Name", anchor=W)
    my_tree.heading("Dose", text="Dose", anchor=CENTER)
    my_tree.heading("Days", text="Days", anchor=W)
    my_tree.heading("Token", text="Token", anchor=W)
    updatetree()

    # Create striped row tags
    my_tree.tag_configure('oddrow', background="white")
    my_tree.tag_configure('evenrow', background="lightblue")


def stock_page():
    s_p = tk.Frame(obj)
    tk.Label(s_p, text='stock Page\n\nPage:3', font='Bold,30').place(x=1,y=1)
def del_page():
    for frame in obj.winfo_children():
        frame.destroy()
    for frame in obj2.winfo_children():
        frame.destroy()
def hideindicate():
    reg_indicate.config(bg=buttoncolorlite)

    stock_indicate.config(bg=buttoncolorlite)

    reg_btn.configure(fg_color=buttoncolorlite)

    stock_btn.configure(fg_color=buttoncolorlite)
def indicate(lb, btn, page):
    hideindicate()
    if btn == 1:
        reg_btn.configure(fg_color=buttoncolor)
    else:
        stock_btn.configure(fg_color=buttoncolor)
    lb.config(bg=buttoncolor)
    del_page()
    page()





# top frames
obj = customtkinter.CTkFrame(master=root, corner_radius=15, width=400, height=600, fg_color=obj_frame_col, border_width=4,
                             border_color="black")
obj.place(x=230, y=130)
obj2 = customtkinter.CTkFrame(master=root, corner_radius=15, width=850, height=600, fg_color=obj_frame_col, border_width=4,
                             border_color="black")
obj2.place(x=650, y=130)


Label(root, text="Clinic Management", width=10, height=2, bg="#c36464", fg='#fff', font='arial 20 bold').pack(side=TOP,
                                                                                                              fill=X)

customtkinter.CTkLabel(master=root, text="Date:", text_color=textcolor,font=(fontmain, 20)).place(x=230, y=77)

# registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
customtkinter.CTkLabel(root, textvariable=Date, text_color=textcolor, font=(fontmain, 20)).place(x=290, y=75)

Date.set(d1)
reg_page()

option_frame = customtkinter.CTkFrame(master=root, corner_radius=0, fg_color=side_frame_col)
my_image = customtkinter.CTkImage(light_image=Image.open("stock-removebg-preview.png"),
                                  dark_image=Image.open("stock-removebg-preview.png"),
                                  size=(40, 40))
reg_btn = customtkinter.CTkButton(option_frame, image=my_image, fg_color=buttoncolor, hover="disable", text='Check-Up',
                                  width=150, corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                                  height=40, command=lambda: indicate(reg_indicate, 1, reg_page))
reg_btn.place(x=15, y=50)

reg_indicate = tk.Label(option_frame, text='', bg=buttoncolor)
reg_indicate.place(x=3, y=55, width=5, height=40)


stock_btn = customtkinter.CTkButton(option_frame, text='Stock            ', hover="disable", image=my_image,
                                    fg_color=buttoncolorlite, width=150, corner_radius=10, border_width=2,
                                    border_color="black", border_spacing=2, height=40,
                                    command=lambda: indicate(stock_indicate, 3, stock_page))
stock_btn.place(x=15, y=160)

stock_indicate = tk.Label(option_frame, text='', bg=buttoncolor)
stock_indicate.place(x=3, y=165, width=5, height=40)

option_frame.pack(side=tk.LEFT)
option_frame.pack_propagate(False)
option_frame.configure(width=200, height=730)
main_frame = tk.Frame(root, highlightbackground='black', highlightthickness=10)

# button

customtkinter.CTkEntry(master=root, corner_radius=15,text_color=textcolor, fg_color=background,textvariable=Search, placeholder_text="search", height=40,
                       font=(fontmain, 20), width=220).place(x=1110, y=75)
imageicon3 = PhotoImage(file="Images/search.png")
srchimage = customtkinter.CTkImage(light_image=Image.open("stock-removebg-preview.png"),
                                   dark_image=Image.open("stock-removebg-preview.png"),
                                   size=(40, 40))
Srch = customtkinter.CTkButton(root, text="Search", command=search, image=srchimage, fg_color=buttoncolor, hover="disable",
                               width=150, corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                               height=40)
Srch.place(x=1350, y=70)

save2Button = customtkinter.CTkButton(obj2, text="Save", image=srchimage, fg_color=buttoncolor, hover="disable", width=150,
                                     corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                                     height=40, command=filldata)
save2Button.place(x=600, y=300)


customtkinter.CTkButton(obj2, text="Reset", image=srchimage, fg_color=buttoncolor, hover="disable", width=150,
                        corner_radius=10, border_width=2, border_color="black", border_spacing=2, height=40,
                        command=Clear2).place(x=600, y=400)
customtkinter.CTkButton(obj2, text="Exit", image=srchimage, fg_color=buttoncolor, hover="disable", width=150,
                        corner_radius=10, border_width=2, border_color="black", border_spacing=2, height=40,
                        command=Exit).place(x=600, y=500)
root.mainloop()