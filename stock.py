from docxtpl import DocxTemplate
from tkcalendar import Calendar, DateEntry
import datetime
from tkinter import *
from datetime import date
from tkinter import messagebox
from PIL import Image
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib
import tkinter as tk
import customtkinter
from tkinter import ttk
import textwrap
from CTkMessagebox import CTkMessagebox
customtkinter.set_appearance_mode("System")

textcolor = "#333333"
side_frame_col = "#F2DFD7"
buttoncolor = "#FF5722"
buttoncolorlite = "#FF855F"
background = "#D4C1EC"
obj_frame_col = "#9F9FED"
mainbackground = "#736ced"
framebg = "#EDEDED"
framefg = "#06283D"
root = customtkinter.CTk()
root.iconbitmap('D:\Programming\Ai and Ml\Machine learning tut\healthcare\stock-removebg-preview.ico')
root.title("Clinic Managment System")
root.state('zoomed')

# root.geometry ("%dx%d"%(width,height))
root.config(bg=mainbackground)
# pathmain = "//LAPTOP-F1A0LRP8/Users/aman/Student_data.xlsx"

stockfiledata = "Stock.xlsx"
stockfile = pathlib.Path(stockfiledata)
if stockfile.exists():
    pass
else:
    stockfile = Workbook()
    sheetstock = stockfile.active
    stockfile.save(stockfiledata)


def Exit():
    root.destroy()


def Clear():
    drug_name.set('')
    drug_salt.set('')
    drug_batch.set('')
    drug_free.set('0')
    drug_mrp.set('')
    drup_pp.set('')
    drug_quan.set('')
def Clear2():

    drug_name.set('')
    drug_salt.set('')
    drug_batch.set('')
    drug_free.set('0')
    drug_mrp.set('')
    drup_pp.set('')
    drug_quan.set('')
def remove_all():
	for record in my_tree.get_children():
		my_tree.delete(record)
def Save():
    file = openpyxl.load_workbook(stockfiledata)
    stockfile1 = file.active
    setdata = StringVar()
    for item in my_tree.get_children():
        values = my_tree.item(item)["values"]
        row = stockfile1.max_row + 1
        for col, value in enumerate(values, start=1):
            setdata.set(value)
            stockfile1.cell(row=row, column=col, value=setdata.get())


    # selected_date = cal.get_date()
    # year = selected_date.year
    # month = selected_date.month
    # day = selected_date.day
    # mainexpdate = datetime.date(year, month, day)
    #
    # daysleft = (mainexpdate - datetime.date.today()).days  # exp date
    # today_main_date = Date.get()  # today date
    # name = drug_name.get()  # name
    # salt = drug_salt.get() # salt
    # batch = drug_batch.get() #batch
    # mrp = drug_mrp.get() # mrp
    # pp = drup_pp.get() # free
    # free = drug_free.get()
    # quan = drug_quan.get()
    # if name == "" or batch == "" or mrp == "" or pp == "" or salt == "" or quan == "":
    #     messagebox.showerror("“error", "Few Data is missing!")
    # else:
    #     file = openpyxl.load_workbook(stockfiledata)
    #     sheet = file.active
    #     sheet.cell(column=1, row=sheet.max_row + 1, value=batch)
    #     sheet.cell(column=2, row=sheet.max_row, value=name)
    #     sheet.cell(column=3, row=sheet.max_row, value=salt)
    #     sheet.cell(column=4, row=sheet.max_row, value=mrp)
    #     sheet.cell(column=5, row=sheet.max_row, value=pp)
    #     sheet.cell(column=6, row=sheet.max_row, value=free)
    #     sheet.cell(column=7, row=sheet.max_row, value=today_main_date)
    #     sheet.cell(column=8, row=sheet.max_row, value=daysleft)
    #     sheet.cell(column=9, row=sheet.max_row, value=quan)
    file.save(stockfiledata)
    remove_all()
        # Clear()  # clear entry box and image section

def search():
    #  # text = Search.get()  # taking input from entry box
    # Clear()
    #
    # file = openpyxl.load_workbook(stockfiledata)
    # sheet = file.active
    # if(text == ""):
    #     messagebox.showerror("Invalid", "Invalid registration number! !!")
    #     return
    # for row in sheet.rows:
    #     if row[0].value == int(text):
    #         name = row[0]
    #         reg_no_position = str(name)[14:-1]
    #         global first
    #         first =1;
    #         reg_number = str(name)[15:-1]
    #         try:
    #             print(str(name))
    #         except:
    #             messagebox.showerror("Invalid", "Invalid registration number! !!")
    # if(first == 0):
    #     messagebox.showerror("Invalid", "Invalid registration number! !!")
    #     return
    # x1 = sheet.cell(row=int(reg_number), column=1).value
    # x2 = sheet.cell(row=int(reg_number), column=2).value
    # x3 = sheet.cell(row=int(reg_number), column=3).value
    # x4 = sheet.cell(row=int(reg_number), column=4).value
    # x5 = sheet.cell(row=int(reg_number), column=5).value
    # x6 = sheet.cell(row=int(reg_number), column=6).value
    # x7 = sheet.cell(row=int(reg_number), column=7).value
    # x8 = sheet.cell(row=int(reg_number), column=8).value
    # x9 = sheet.cell(row=int(reg_number), column=9).value
    # x10 = sheet.cell(row=int(reg_number), column=10).value
    # x11 = sheet.cell(row=int(reg_number), column=11).value
    # x12 = sheet.cell(row=int(reg_number), column=12).value
    # x13 = sheet.cell(row=int(reg_number), column=13).value
    # x14 = sheet.cell(row=int(reg_number), column=14).value
    #
    # Registration.set(x1)
    # Name.set(x2)
    #
    #
    #
    #
    # age.set(x5)
    #
    # Date.set(x6)
    #
    # weight.set(x7)
    #
    # height.set(x8)
    # temprature.set(x9)
    #
    # respiration.set(x10)
    # pulse.set(x11)
    # bp.set(x12)
    # village.set(x13)
    # mobile.set(x14)
    # first =0;
    pass
    ####################################Update#####################################


def Update():
    # reg_number = IntVar()
    # batch = drug_batch.get()
    # name = drug_name.get()
    # salt = drug_salt.get()
    # today_date = Date.get()
    # mrp = drug_mrp.get()
    # pp = drup_pp.get()
    # free = drug_free.get()
    # file = openpyxl.load_workbook(stockfiledata)
    # sheet = file.active
    # for row in sheet.rows:
    #     if row[0].value == R1:
    #         name = row[0]
    #         print(str(name))
    #         reg_no_position = str(name)[14:-1]
    #         reg_number = str(name)[15:-1]
    #         print(reg_number)
    #
    # sheet.cell(column=1, row=int(reg_number), value=R1)
    # sheet.cell(column=2, row=int(reg_number), value=N1)
    # sheet.cell(column=5, row=int(reg_number), value=D2)
    # sheet.cell(column=6, row=int(reg_number), value=D1)
    # sheet.cell(column=7, row=int(reg_number), value=Re1)
    # sheet.cell(column=8, row=int(reg_number), value=S1)
    # sheet.cell(column=9, row=int(reg_number), value=fathername)
    # sheet.cell(column=10, row=int(reg_number), value=mothername)
    # sheet.cell(column=11, row=int(reg_number), value=F1)
    # sheet.cell(column=12, row=int(reg_number), value=M1)
    # sheet.cell(column=13, row=int(reg_number), value=uvill)
    # sheet.cell(column=14, row=int(reg_number), value=umob)
    # file.save(pathmain)
    # messagebox.showinfo("info", "Sucessfully data entered!!!")
    # Clear()
    pass
    ####################################gender#####################################



patient_name = StringVar()
docter_name = StringVar()


fontmain = "Dotum"
drug_name = StringVar()
drug_salt= StringVar()

drug_mrp = StringVar()
drup_pp = StringVar()
drug_free = StringVar()
drug_batch = StringVar()
drug_quan = StringVar()
drug_hns = StringVar()
drug_discount = StringVar()
drug_sgst = StringVar()
drug_cgst = StringVar()
drug_amount = StringVar()
drug_form = StringVar()
drug_gst = StringVar()
drug_date = StringVar()
obj = None

expdate = StringVar()


quan_bill = StringVar()
med_name = StringVar()
Search = StringVar()
data = []
maintainstock = []


# top frames
obj = customtkinter.CTkFrame(master=root, corner_radius=15, width=1250, height=650, fg_color=obj_frame_col, border_width=4,
                             border_color="black",bg_color=mainbackground)
obj.place(x=230, y=100)

Label(root, text="Clinic Management", width=10, height=2, bg="#c36464", fg='#fff', font='arial 20 bold').pack(side=TOP,
                                                                                                              fill=X)


Registration = IntVar()
Date = StringVar()


today = date.today()
d1 = today.strftime("%d/%m/%Y")

Date.set(d1)

# def updatetree():
#
#     file = openpyxl.load_workbook(patient_detail_from_doc)
#     # sheet.cell(column=1, row=sheet.max_row + 1, value=med_no)
#     sheet1 = file.active
#
#     for row in sheet1.rows:
#         lst = []
#         if Search.get() == row[0].value:
#             lst.append(row[0].value)
#             lst.append(row[1].value)
#             lst.append(row[2].value)
#             lst.append(row[3].value)
#             data.append(lst)
#
#     global count
#     count = 0
#
#     for record in data:
#         if count % 2 == 0:
#             my_tree.insert(parent='', index='end', iid=count, text="", values=(record[0], record[1], record[2],record[3]),
#                            tags=('evenrow',))
#         else:
#             my_tree.insert(parent='', index='end', iid=count, text="", values=(record[0], record[1], record[2],record[3]),
#                            tags=('oddrow',))
#
#         count += 1
def filldata():
    file = openpyxl.load_workbook(stockfiledata)
    sheet5 = file.active
    setdata = StringVar()
    for item in my_tree.get_children():
        values = my_tree.item(item)["values"]
        row = sheet5.max_row + 1
        for col, value in enumerate(values, start=1):
            setdata.set(value)
            sheet5.cell(row=row, column=col, value=setdata.get())
    file.save(stockfiledata)
    file.close()
def wrap(string, lenght=8):
    return '\n'.join(textwrap.wrap(string, lenght))
def gendays():
    selected_date = cal.get_date()
    year = selected_date.year
    month = selected_date.month
    day = selected_date.day
    mainexpdate = datetime.date(year, month, day)
    daysleft = (mainexpdate - datetime.date.today()).days
    return daysleft
count = 0
def remove_all():
	for record in my_tree.get_children():
		my_tree.delete(record)

def reg_page():

    drug_free.set('0')
    customtkinter.CTkLabel(obj, text="BATCH:", text_color=textcolor, font=(fontmain, 20)).place(x=30, y=55)
    customtkinter.CTkLabel(obj, text="DRUG NAME:", text_color=textcolor, font=(fontmain, 20)).place(x=30, y=105)
    customtkinter.CTkLabel(obj, text="PP:", text_color=textcolor, font=(fontmain, 20)).place(x=30, y=155)
    customtkinter.CTkLabel(obj, text="Free:", text_color=textcolor, font=(fontmain, 20)).place(x=30, y=205)

    customtkinter.CTkLabel(obj, text="SALT NAME:", text_color=textcolor, font=(fontmain, 20)).place(x=450, y=55)
    customtkinter.CTkLabel(obj, text="Quantity:", text_color=textcolor, font=(fontmain, 20)).place(x=450, y=105)
    customtkinter.CTkLabel(obj, text="MRP:", text_color=textcolor, font=(fontmain, 20)).place(x=450, y=155)
    customtkinter.CTkLabel(obj, text="EXPIRE DATE:", text_color=textcolor, font=(fontmain, 20)).place(x=450, y=205)
    customtkinter.CTkLabel(obj, text="FORM NAME:", text_color=textcolor, font=(fontmain, 20)).place(x=450, y=255)

    customtkinter.CTkLabel(obj, text="HSN Code:", text_color=textcolor, font=(fontmain, 20)).place(x=870, y=55)
    customtkinter.CTkLabel(obj, text="Discount:", text_color=textcolor, font=(fontmain, 20)).place(x=870, y=105)
    customtkinter.CTkLabel(obj, text="SGST:", text_color=textcolor, font=(fontmain, 20)).place(x=870, y=155)
    customtkinter.CTkLabel(obj, text="CGST:", text_color=textcolor, font=(fontmain, 20)).place(x=870, y=205)
    customtkinter.CTkLabel(obj, text="FORM GST:", text_color=textcolor, font=(fontmain, 20)).place(x=870, y=255)

    # Entry
    batch_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                        textvariable=drug_batch, height=40, font=(fontmain, 20),
                                        width=220)
    batch_entry.place(x=170, y=50)
    name_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                       textvariable=drug_name, height=40, font=(fontmain, 20),
                                       width=220)
    name_entry.place(x=170, y=100)
    pp_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                      textvariable=drup_pp, height=40,
                                      font=(fontmain, 20), width=220)
    pp_entry.place(x=170, y=150)
    free_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                        textvariable=drug_free, height=40,
                                        font=(fontmain, 20), width=220)
    free_entry.place(x=170, y=200)



    salt_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                          textvariable=drug_salt, height=40,
                                          font=(fontmain, 20), width=220)
    salt_entry.place(x=600, y=50)
    quan_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                        textvariable=drug_quan, height=40,
                                        font=(fontmain, 20), width=220)
    quan_entry.place(x=600, y=100)
    mrp_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                          textvariable=drug_mrp, height=40,
                                          font=(fontmain, 20), width=220)
    mrp_entry.place(x=600, y=150)
    global cal
    cal = DateEntry(obj, width=22,
                    foreground='yellow', year=2023, tooltipbackground="yellow", font=(fontmain, 13))
    cal.place(x=600, y=200)
    form_name_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                            textvariable=drug_form, height=40,
                                            font=(fontmain, 20), width=220)
    form_name_entry.place(x=600, y=250)


    hsn_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                        textvariable=drug_hns, height=40, font=(fontmain, 20),
                                        width=220)
    hsn_entry.place(x=1000, y=50)
    disc_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                       textvariable=drug_discount, height=40, font=(fontmain, 20),
                                       width=220)
    disc_entry.place(x=1000, y=100)
    sgst_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                      textvariable=drug_sgst, height=40,
                                      font=(fontmain, 20), width=220)
    sgst_entry.place(x=1000, y=150)
    cgst_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                        textvariable=drug_cgst, height=40,
                                        font=(fontmain, 20), width=220)
    cgst_entry.place(x=1000, y=200)
    form_gst_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor, fg_color=background, corner_radius=15,
                                        textvariable=drug_gst, height=40,
                                        font=(fontmain, 20), width=220)
    form_gst_entry.place(x=1000, y=250)

    # def Clear():
    #     global pnt_dis
    #     med_no.set('')
    #     days.set('')
    #     dose.set('')
    #     pnt_dis = ''
    #     med_name.set('')
    #     # registration_no()
     # exp date
    def add_record():
        daysleft = gendays()
        getamount = float(drup_pp.get())*int(drug_quan.get())
        gettodaydate = datetime.date.today()
        drug_date.set(gettodaydate)
        formatted_number = f"{getamount:.3f}"
        drug_amount.set(str(formatted_number))
        my_tree.tag_configure('oddrow', background="white")
        my_tree.tag_configure('evenrow', background="lightblue")

        global count
        if count % 2 == 0:
            my_tree.insert(parent='', index='end', iid=count, text="",
                           values=(drug_batch.get(), drug_name.get(), drug_salt.get(), drug_mrp.get(),drup_pp.get(),drug_free.get(),daysleft,drug_quan.get(),drug_hns.get(),drug_discount.get(),drug_sgst.get(),drug_cgst.get(),drug_amount.get(),drug_form.get(),drug_gst.get(),drug_date.get()), tags=('evenrow',))
        else:
            my_tree.insert(parent='', index='end', iid=count, text="",
                           values=(drug_batch.get(), drug_name.get(), drug_salt.get(), drug_mrp.get(),drup_pp.get(),drug_free.get(),daysleft,drug_quan.get(),drug_hns.get(),drug_discount.get(),drug_sgst.get(),drug_cgst.get(),drug_amount.get(),drug_form.get(),drug_gst.get(),drug_date.get()), tags=('oddrow',))

        count += 1

    def find_item_in_list(lst, item):
        try:
            index = lst.index(item)
            return False
        except ValueError:
            return True



    def delmed():
        x = my_tree.selection()[0]
        my_tree.delete(x)

    add_btn = customtkinter.CTkButton(obj, text='ADD', hover="disable",
                                      fg_color=buttoncolorlite, width=80, corner_radius=10, border_width=2,
                                      border_color="black", border_spacing=2, height=40, command=lambda: add_record()
                                      )
    add_btn.place(x=30, y=280)
    delete_btn = customtkinter.CTkButton(obj, text='DELETE', hover="disable",
                                         fg_color=buttoncolorlite, width=100, corner_radius=10, border_width=2,
                                         border_color="black", border_spacing=2, height=40,
                                         command=lambda: delmed()
                                         )
    delete_btn.place(x=140, y=280)
    #     excel

    style = ttk.Style()
    # Pick a theme
    style.theme_use("clam")
    # Configure our treeview colors

    style.configure("Treeview",
                    background=background,
                    foreground="black",
                    rowheight=22,
                    fieldbackground=mainbackground
                    )
    # Change selected color
    style.map('Treeview',
              background=[('selected', 'blue')])

    # Create Treeview Frame
    tree_frame = Frame(obj)
    tree_frame.place(x=20, y=340,height=230,width=1200)

    # Treeview Scrollbar

    scroll_x = customtkinter.CTkScrollbar(tree_frame,orientation=HORIZONTAL,width=1200,height = 22,corner_radius=9, fg_color=mainbackground,
                                          button_color=background)
    scroll_y = customtkinter.CTkScrollbar(tree_frame, orientation=VERTICAL,width=22,height = 210,corner_radius=9, fg_color=mainbackground,
                                       button_color=background)
    # tree_scroll = customtkinter.CTkScrollbar(tree_frame, corner_radius=9, fg_color=mainbackground,
    #                                          button_color=background)
    # tree_scroll.pack(side=RIGHT, fill=Y)


    # Create Treeview
    global my_tree
    main_tree_frame = Frame(tree_frame)
    main_tree_frame.place(x=0, y=0, height=208, width=1178)
    my_tree = ttk.Treeview(main_tree_frame)
    # Pack to the screen
    my_tree.pack()
    my_tree.configure(yscrollcommand=scroll_y.set,xscrollcommand=scroll_x.set)
    my_tree.configure(selectmode="extended")
    # Configure the scrollbar
    scroll_x.configure(command=my_tree.xview)
    scroll_y.configure(command=my_tree.yview)
    # tree_scroll.configure(command=my_tree.yview)
    scroll_x.place(x=0,y=208)
    scroll_y.place(x=1178,y=0)
    # Define Our Columns
    my_tree.configure(columns=("Batch", "Name", "Salt", "Mrp", "PP", "Free","Days","Quantity","HSN","Discount","SGST","CGST","Amount","Form Name","Form Number","Date"))
    # my_tree['columns'] = ("Batch", "Name", "Salt", "Mrp", "PP", "Free","Days","Quantity","HSN","Discount","SGST","CGST","Amount")

    # Formate Our Columns
    my_tree.column("#0", width=0, stretch=NO)
    my_tree.column("Batch", anchor=W, width=120)
    my_tree.column("Name", anchor=W, width=220)
    my_tree.column("Salt", anchor=W, width=220)
    my_tree.column("Mrp", anchor=W, width=60)
    my_tree.column("PP", anchor=W, width=60)
    my_tree.column("Free", anchor=W, width=60)
    my_tree.column("Days", anchor=W, width=60)
    my_tree.column("Quantity", anchor=W, width=60)
    my_tree.column("HSN", anchor=W, width=60)
    my_tree.column("Discount", anchor=W, width=60)
    my_tree.column("SGST", anchor=W, width=60)
    my_tree.column("CGST", anchor=W, width=60)
    my_tree.column("Amount", anchor=W, width=60)
    my_tree.column("Form Name", anchor=W, width=220)
    my_tree.column("Form Number", anchor=W, width=200)
    my_tree.column("Date", anchor=W, width=80)

    # Create Headings
    my_tree.heading("#0", text="", anchor=W)
    my_tree.heading("Batch", text="Batch", anchor=W)
    my_tree.heading("Name", text="Name", anchor=W)
    my_tree.heading("Salt", text="Salt", anchor=W)
    my_tree.heading("Mrp", text="Mrp", anchor=W)
    my_tree.heading("PP", text="PP", anchor=W)
    my_tree.heading("Free", text="Free", anchor=W)
    my_tree.heading("Days", text="Days", anchor=W)
    my_tree.heading("Quantity", text="Quantity", anchor=W)
    my_tree.heading("HSN", text="HSN", anchor=W)
    my_tree.heading("Discount", text="Discount", anchor=W)
    my_tree.heading("SGST", text="SGST", anchor=W)
    my_tree.heading("CGST", text="CGST", anchor=W)
    my_tree.heading("Amount", text="Amount", anchor=W)
    my_tree.heading("Form Name", text="Form Name", anchor=W)
    my_tree.heading("Form Number", text="Form Number", anchor=W)
    my_tree.heading("Date", text="Date", anchor=W)


    # updatetree()

    # Create striped row tags
    my_tree.tag_configure('oddrow', background="white")
    my_tree.tag_configure('evenrow', background="lightblue")
    saveButton = customtkinter.CTkButton(obj, text="Save", image=srchimage, fg_color=buttoncolor, hover="disable",
                                         width=150,
                                         corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                                         height=40, command=Save)
    saveButton.place(x=50, y=580)
    customtkinter.CTkButton(obj, text="Reset", image=srchimage, fg_color=buttoncolor, hover="disable", width=150,
                            corner_radius=10, border_width=2, border_color="black", border_spacing=2, height=40,
                            command=Clear2).place(x=250, y=580)

def remove_all():
	for record in my_tree.get_children():
		my_tree.delete(record)
def search_patient():

    text = Search.get()

    # Clear()

    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            try:
                print(str(name))
            except:
                messagebox.showerror("Invalid", "Invalid registration number! !!")

    x1 = sheet.cell(row=int(reg_number), column=2).value
    x2 = sheet.cell(row=int(reg_number), column=2).value

    patient_name.set(x1)
    docter_name.set(x2)
    updatetree()
def updatetree():

    remove_all()
    data.clear()

    file = openpyxl.load_workbook("patient.xlsx")
    # sheet.cell(column=1, row=sheet.max_row + 1, value=med_no)
    sheet1 = file.active

    for row in sheet1.rows:
        lst = []
        if Search.get() == row[0].value:
            lst.append(row[0].value)
            lst.append(row[1].value)
            lst.append(row[2].value)
            lst.append(row[3].value)
            data.append(lst)

    global count
    count = 0

    for record in data:
        if count % 2 == 0:
            my_tree.insert(parent='', index='end', iid=count, text="", values=(record[0], record[1], record[2],record[3]),
                           tags=('evenrow',))
        else:
            my_tree.insert(parent='', index='end', iid=count, text="", values=(record[0], record[1], record[2],record[3]),
                           tags=('oddrow',))

        count += 1

def get_row_values(file_path,row_number, column_indices=None, column_names=None):
    file = openpyxl.load_workbook("Stock.xlsx")
    sheet = file.active
    row_values = []
    row_values = []
    for col_index in column_indices:
        cell = sheet.cell(row=row_number, column=col_index)
        row_values.append(cell.value)

    file.close()

    return row_values
def bill_page():
    def get_treeview_data(treeview):
        data = []
        count = 1
        for item in treeview.get_children():
            values = treeview.item(item)['values']
            values.insert(0, count)  # Insert the counter at the beginning of the values list
            data.append(values)
            count += 1
        return data


    customtkinter.CTkEntry(master=obj, corner_radius=15, text_color=textcolor, fg_color=background,
                           textvariable=Search, placeholder_text="search", height=40,
                           font=(fontmain, 20), width=220).place(x=700, y=270)
    customtkinter.CTkLabel(obj, text="Quantity", text_color=textcolor, font=(fontmain, 20)).place(x=1000, y=120)
    customtkinter.CTkEntry(master=obj, corner_radius=15, text_color=textcolor, fg_color=background,
                           textvariable=quan_bill, height=40,
                           font=(fontmain, 20), width=220).place(x=1000, y=170)
    Srch = customtkinter.CTkButton(obj, text="Search", command=search_patient, image=srchimage, fg_color=buttoncolor,
                                   hover="disable",
                                   width=150, corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                                   height=40)
    Srch.place(x=1000, y=270)

    def update(data):
        # Clear the listbox
        my_list.delete(0, END)

        # Add toppings to listbox
        for item in data:
            my_list.insert(END, item)
    def fillout(e):
        my_entry.delete(0, END)

        selected_item = my_list.get(ANCHOR)
        my_entry.insert(0, selected_item)

        my_list.selection_clear(0, END)
    def check(e):
        # grab what was typed
        typed = my_entry.get()

        if typed == '':
            data = toppings
        else:
            data = []
            for item in toppings:
                if typed.lower() in item.lower():
                    data.append(item)

        # update our listbox with selected items
        update(data)

    customtkinter.CTkLabel(obj, text="Search", text_color=textcolor, font=(fontmain, 20)).place(x=700, y=10)
    my_entry = customtkinter.CTkEntry(master=obj, text_color=textcolor,fg_color=background,corner_radius=10, textvariable=med_name, height=30,
                                        font=(fontmain, 20), width=240)
    my_entry.place(x=700, y=50)

    my_list = Listbox(obj, width=40,bd=0,background=background)
    my_list.place(x=700, y=90)

    # Create a list of pizza toppings
    toppings = []
    file = openpyxl.load_workbook("Stock.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if(row[1].value != None):
            toppings.append(row[1].value)

    # Add the toppings to our list
    update(toppings)

    # Create a binding on the listbox onclick
    my_list.bind("<<ListboxSelect>>", fillout)

    # Create a binding on the entry box
    my_entry.bind("<KeyRelease>", check)
    style = ttk.Style()
    # Pick a theme
    style.theme_use("clam")
    # Configure our treeview colors

    style.configure("Treeview",
                    background=background,
                    foreground="black",
                    rowheight=20,
                    fieldbackground=mainbackground
                    )
    # Change selected color
    style.map('Treeview',
              background=[('selected', 'blue')])

    # Create Treeview Frame
    tree_frame = Frame(obj)
    tree_frame.place(x=20, y=20)

    # Treeview Scrollbar
    tree_scroll = customtkinter.CTkScrollbar(tree_frame, corner_radius=9, fg_color=mainbackground,
                                             button_color=background)
    tree_scroll.pack(side=RIGHT, fill=Y)

    # Create Treeview
    global my_tree
    my_tree = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set, selectmode="extended")
    # Pack to the screen
    my_tree.pack()

    # Configure the scrollbar
    tree_scroll.configure(command=my_tree.yview)

    # Define Our Columns
    my_tree['columns'] = ("Token", "Name", "Dose", "Days")

    # Formate Our Columns
    my_tree.column("#0", width=0, stretch=NO)
    my_tree.column("Name", anchor=W, width=140)
    my_tree.column("Dose", anchor=CENTER, width=100)
    my_tree.column("Days", anchor=W, width=140)
    my_tree.column("Token", anchor=W, width=140)

    # Create Headings
    my_tree.heading("#0", text="", anchor=W)
    my_tree.heading("Name", text="Name", anchor=W)
    my_tree.heading("Dose", text="Dose", anchor=CENTER)
    my_tree.heading("Days", text="Days", anchor=W)
    my_tree.heading("Token", text="Token", anchor=W)
    updatetree()

    # Create striped row tags
    my_tree.tag_configure('oddrow', background="white")
    my_tree.tag_configure('evenrow', background="lightblue")


    def add_record():
        compare_value = my_entry.get()
        file = openpyxl.load_workbook("Stock.xlsx")
        sheet = file.active
        flag = 0
        row_num =0
        for row in sheet.rows:
            row_num = row_num +1
            if row[1].value == compare_value:
                if int(row[7].value) < int(quan_bill.get()) or int(row[6].value) <= 0:
                    flag =1
                else:
                    flag =0
                break

        if(flag == 0):
            lst_stock = []
            file = openpyxl.load_workbook("Stock.xlsx")
            sheet = file.active
            main_file_name = "Stock.xlsx"
            row_values = []
            columns = [2,1,7,5,4,8,11,12]
            row_values = get_row_values(main_file_name, row_num, column_indices=columns)
            print(row_values)
            final_val = int(row_values[5])-int(quan_bill.get())
            sheet.cell(column=8, row=int(row_num), value=final_val)
            total_amount = int(quan_bill.get()) * float(row_values[4])
            my_tree2.tag_configure('oddrow', background="white")
            my_tree2.tag_configure('evenrow', background="lightblue")
            file.save("Stock.xlsx")
            global count
            if count % 2 == 0:
                my_tree2.insert(parent='', index='end', iid=count, text="",
                                values=(
                                    row_values[0], row_values[1], row_values[2], row_values[3], row_values[4],
                                    quan_bill.get(), row_values[5], row_values[6],total_amount), tags=('evenrow'))
            else:
                my_tree2.insert(parent='', index='end', iid=count, text="",
                                values=(
                                    row_values[0], row_values[1], row_values[2], row_values[3], row_values[4],
                                    quan_bill.get(), row_values[5], row_values[6], total_amount), tags=('oddrow'))

            count += 1
        else:
            CTkMessagebox(message="CTkMessagebox is successfully installed.",
                          icon="check", option_1="Thanks")



    def find_item_in_list(lst, item):
        try:
            index = lst.index(item)
            return False
        except ValueError:
            return True

    #     x = my_tree2.selection()[0]
    def delmed():
        selection = my_tree2.selection()
        if selection:
            x = selection[0]
            values = my_tree2.item(x)['values']
            my_tree2.delete(x)

            # Open the Excel file
            file_path = "Stock.xlsx"
            file = openpyxl.load_workbook(file_path)

            # Get the active sheet
            sheet = file.active
            # for row in sheet.rows:
            #     row_num = row_num + 1
            #     if row[1].value == compare_value:
            #         if int(row[7].value) < int(quan_bill.get()) or int(row[6].value) <= 0:
            #             flag = 1
            #         else:
            #             flag = 0
            #         break
            # Find the first element of the value in the Excel sheet
            row_num = 0
            for row in sheet.rows:
                row_num = row_num + 1
                if row[1].value == values[0]:
                    cell = sheet.cell(row=row_num, column=8)
                    cell.value = str(int(cell.value) + int(values[5]))
                    sheet.cell(row=row_num, column=8,value = cell.value)

            # Save the changes
            file.save(file_path)
            file.close()

            return values
        else:
            return None
    # def delmed():
    #     x = my_tree2.selection()[0]
    #     my_tree2.delete(x)

    add_btn = customtkinter.CTkButton(obj, text='ADD', hover="disable",
                                      fg_color=buttoncolorlite, width=80, corner_radius=10, border_width=2,
                                      border_color="black", border_spacing=2, height=40, command=lambda: add_record()
                                      )
    add_btn.place(x=30, y=280)
    delete_btn = customtkinter.CTkButton(obj, text='DELETE', hover="disable",
                                         fg_color=buttoncolorlite, width=100, corner_radius=10, border_width=2,
                                         border_color="black", border_spacing=2, height=40,
                                         command=lambda: delmed()
                                         )
    delete_btn.place(x=140, y=280)
    #     excel

    style = ttk.Style()
    # Pick a theme
    style.theme_use("clam")
    # Configure our treeview colors

    style.configure("Treeview",
                    background=background,
                    foreground="black",
                    rowheight=22,
                    fieldbackground=mainbackground
                    )
    # Change selected color
    style.map('Treeview',
              background=[('selected', 'blue')])

    # Create Treeview Frame
    tree_frame = Frame(obj)
    tree_frame.place(x=20, y=540,height=230,width=785)

    # Treeview Scrollbar

    scroll_x = customtkinter.CTkScrollbar(tree_frame,orientation=HORIZONTAL,width=785,height = 22,corner_radius=9, fg_color=mainbackground,
                                          button_color=background)
    scroll_y = customtkinter.CTkScrollbar(tree_frame, orientation=VERTICAL,width=22,height = 210,corner_radius=9, fg_color=mainbackground,
                                       button_color=background)
    # tree_scroll = customtkinter.CTkScrollbar(tree_frame, corner_radius=9, fg_color=mainbackground,
    #                                          button_color=background)
    # tree_scroll.pack(side=RIGHT, fill=Y)


    # Create Treeview
    global my_tree2
    main_tree_frame = Frame(tree_frame)
    main_tree_frame.place(x=0, y=0, height=208, width=763)
    my_tree2 = ttk.Treeview(main_tree_frame)
    # Pack to the screen
    my_tree2.pack()
    my_tree2.configure(yscrollcommand=scroll_y.set,xscrollcommand=scroll_x.set)
    my_tree2.configure(selectmode="extended")
    # Configure the scrollbar
    scroll_x.configure(command=my_tree2.xview)
    scroll_y.configure(command=my_tree2.yview)
    # tree_scroll.configure(command=my_tree.yview)
    scroll_x.place(x=0,y=208)
    scroll_y.place(x=763,y=0)
    # Define Our Columns
    my_tree2.configure(columns=("Name","Batch","Days", "PP", "Mrp","Quantity","SGST","CGST","Amount"))
    # my_tree['columns'] = ("Batch", "Name", "Salt", "Mrp", "PP", "Free","Days","Quantity","HSN","Discount","SGST","CGST","Amount")

    # Formate Our Columns
    my_tree2.column("#0", width=0, stretch=NO)
    my_tree2.column("Name", anchor=W, width=220)
    my_tree2.column("Batch", anchor=W, width=120)
    my_tree2.column("Days", anchor=W, width=60)
    my_tree2.column("PP", anchor=W, width=60)
    my_tree2.column("Mrp", anchor=W, width=60)
    my_tree2.column("Quantity", anchor=W, width=60)
    my_tree2.column("SGST", anchor=W, width=60)
    my_tree2.column("CGST", anchor=W, width=60)
    my_tree2.column("Amount", anchor=W, width=60)

    # my_tree2.column("Salt", anchor=W, width=220)#-------------------
    # my_tree2.column("Free", anchor=W, width=60)#-------------------
    # my_tree2.column("HSN", anchor=W, width=60)#-------------------
    # my_tree2.column("Discount", anchor=W, width=60)#-------------------
    # my_tree2.column("Form Name", anchor=W, width=220)#-------------------
    # my_tree2.column("Form Number", anchor=W, width=200)#-------------------
    # my_tree2.column("Date", anchor=W, width=80)#-------------------

    # Create Headings
    my_tree2.heading("#0", text="", anchor=W)
    my_tree2.heading("Name", text="Name", anchor=W)
    my_tree2.heading("Batch", text="Batch", anchor=W)
    my_tree2.heading("Days", text="Days", anchor=W)
    my_tree2.heading("PP", text="PP", anchor=W)
    my_tree2.heading("Mrp", text="Mrp", anchor=W)
    my_tree2.heading("Quantity", text="Quantity", anchor=W)
    my_tree2.heading("SGST", text="SGST", anchor=W)
    my_tree2.heading("CGST", text="CGST", anchor=W)
    my_tree2.heading("Amount", text="Amount", anchor=W)


    # my_tree2.heading("Salt", text="Salt", anchor=W) #-------------------
    # my_tree2.heading("Free", text="Free", anchor=W)#-------------------
    # my_tree2.heading("HSN", text="HSN", anchor=W)#-------------------
    # my_tree2.heading("Discount", text="Discount", anchor=W)#-------------------
    # my_tree2.heading("Form Name", text="Form Name", anchor=W)#-------------------
    # my_tree2.heading("Form Number", text="Form Number", anchor=W)#-------------------
    # my_tree2.heading("Date", text="Date", anchor=W)#-------------------
    def save_the_bill():
        doc = DocxTemplate("medical_bill_tamplet.docx")
        name = Search.get()+patient_name.get()+docter_name.get()
        # phone = phone_entry.get()
        invoice_list  = get_treeview_data(my_tree2)
        print(invoice_list)
        subtotal = sum(float(item[3]) for item in invoice_list)
        salestax = 0.1
        total = subtotal * (1 - salestax)

        doc.render({"patientname": patient_name.get(),
                    "docname": docter_name.get(),
                    "date": datetime.date.today(),
                    "sno": "1",
                    "invoice_list": invoice_list,
                    "grandtotal": str(salestax * 100) + "%",
                    "total": subtotal})

        doc_name = "new_invoice" + name + datetime.datetime.now().strftime("%Y-%m-%d-%H%M%S") + ".docx"
        doc.save(doc_name)

        messagebox.showinfo("Invoice Complete", "Invoice Complete")
    save_bill = customtkinter.CTkButton(obj, text="Save", command=save_the_bill, image=srchimage, fg_color=buttoncolor,
                                   hover="disable",
                                   width=150, corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                                   height=40)
    save_bill.place(x=1000, y=590)
def del_page():
    for frame in obj.winfo_children():
        frame.destroy()


def hideindicate():
    reg_indicate.config(bg=buttoncolorlite)
    bill_indicate.config(bg=buttoncolorlite)
    stock_indicate.config(bg=buttoncolorlite)

    reg_btn.configure(fg_color=buttoncolorlite)
    bill_btn.configure(fg_color=buttoncolorlite)
    # stock_btn.configure(fg_color=buttoncolorlite)


def indicate(lb, btn, page):
    hideindicate()
    if btn == 1:
        reg_btn.configure(fg_color=buttoncolor)
    elif btn == 2:
        bill_btn.configure(fg_color=buttoncolor)
    # else:
    #     stock_btn.configure(fg_color=buttoncolor)
    lb.config(bg=buttoncolor)
    del_page()
    page()


option_frame = customtkinter.CTkFrame(master=root, corner_radius=0, fg_color=side_frame_col)
my_image = customtkinter.CTkImage(light_image=Image.open("stock-removebg-preview.png"),
                                  dark_image=Image.open("stock-removebg-preview.png"),
                                  size=(40, 40))
reg_btn = customtkinter.CTkButton(option_frame, image=my_image, fg_color=buttoncolorlite, hover="disable", text='Stock Update',
                                  width=150, corner_radius=10, border_width=2, border_color="black", border_spacing=2,
                                  height=40, command=lambda: indicate(reg_indicate, 1, reg_page))
reg_btn.place(x=15, y=50)

reg_indicate = tk.Label(option_frame, text='', bg=buttoncolor)
reg_indicate.place(x=3, y=55, width=5, height=40)

bill_btn = customtkinter.CTkButton(option_frame, text="Bill                 ", hover="disable", image=my_image,
                                   fg_color=buttoncolorlite, width=150, corner_radius=10, border_width=2, border_color="black",
                                   border_spacing=2, height=40, command=lambda: indicate(bill_indicate, 2, bill_page))
bill_btn.place(x=15, y=105)

bill_indicate = tk.Label(option_frame, text='', bg=buttoncolor)
bill_indicate.place(x=3, y=110, width=5, height=40)

# stock_btn = customtkinter.CTkButton(option_frame, text='Stock            ', hover="disable", image=my_image,
#                                     fg_color=buttoncolorlite, width=150, corner_radius=10, border_width=2,
#                                     border_color="black", border_spacing=2, height=40,
#                                     command=lambda: indicate(stock_indicate, 3, stock_page))
# stock_btn.place(x=15, y=160)

stock_indicate = tk.Label(option_frame, text='', bg=buttoncolor)
stock_indicate.place(x=3, y=165, width=5, height=40)

option_frame.pack(side=tk.LEFT)
option_frame.pack_propagate(False)
option_frame.configure(width=200, height=730)
main_frame = tk.Frame(root, highlightbackground='black', highlightthickness=10)

# button


# customtkinter.CTkEntry(master=root, corner_radius=15,text_color=textcolor, fg_color=background,textvariable=Search, placeholder_text="search", height=40,
#                        font=(fontmain, 20), width=220).place(x=1110, y=75)
# imageicon3 = PhotoImage(file="Images/search.png")
srchimage = customtkinter.CTkImage(light_image=Image.open("stock-removebg-preview.png"),
                                   dark_image=Image.open("stock-removebg-preview.png"),
                                   size=(40, 40))
# Srch = customtkinter.CTkButton(root, text="Search", command=search, image=srchimage, fg_color=buttoncolor, hover="disable",
#                                width=150, corner_radius=10, border_width=2, border_color="black", border_spacing=2,
#                                height=40)
# Srch.place(x=1350, y=70)

# customtkinter.CTkButton(root, text="Upload", image=srchimage, fg_color=buttoncolor, hover="disable", width=150,
#                         corner_radius=10, border_width=2, border_color="black", border_spacing=2, height=40,
#                         command=Update).place(x=1300, y=370)

# customtkinter.CTkButton(root, text="Exit", image=srchimage, fg_color=buttoncolor, hover="disable", width=150,
#                         corner_radius=10, border_width=2, border_color="black", border_spacing=2, height=40,
#                         command=Exit).place(x=1300, y=610)

root.mainloop()